"""Comprehensive spreadsheet manager combining parsing and data management functionality."""

import openpyxl
import numpy as np
import pandas as pd
import os, csv, re, logging, json, uuid
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any, Optional, Union, Tuple
from openpyxl.worksheet.datavalidation import DataValidation

class SpreadsheetManager:
    """Comprehensive spreadsheet manager handling file parsing, data management, and operations."""
    
    def __init__(self, file_id: str = None, original_filename: str = None, data_df: Optional[pd.DataFrame] = None, file_path: Optional[str] = None, original_file_type: Optional[str] = None):
        """Initialize spreadsheet manager with optional data."""
        self.logger = logging.getLogger(__name__)
        self.header_patterns = [
            r'^(total|sum|subtotal|grand total|net|gross)$',
            r'^(header|title|label)$',
            r'^\s*$',
            r'^[A-Z\s]+$',
        ]
        
        # Data management properties
        self.file_id = file_id or str(uuid.uuid4())
        self.original_filename = original_filename
        self.data_df = data_df
        self.file_path = file_path
        self.original_file_type = original_file_type or (os.path.splitext(original_filename)[1].lower() if original_filename else None)
        
        if data_df is not None:
            column_names = [f"Column_{i}" for i in range(len(data_df.columns))]
        else:
            column_names = []
            
        self.metadata = {
            'filename': original_filename,
            'columns': column_names,
            'rows': len(data_df) if data_df is not None else 0,
            'is_preprocessed': True,
            'original_file_type': self.original_file_type
        }
    
    @classmethod
    def from_file(cls, file_path: str, file_id: str = None) -> 'SpreadsheetManager':
        """Create SpreadsheetManager by parsing a file."""
        manager = cls()
        df, sheets, original_file_type = manager.parse_file(file_path)
        
        filename = os.path.basename(file_path)
        manager.file_id = file_id or str(uuid.uuid4())
        manager.original_filename = filename
        manager.data_df = df
        manager.file_path = file_path
        manager.original_file_type = original_file_type
        
        if df is not None:
            column_names = [f"Column_{i}" for i in range(len(df.columns))]
        else:
            column_names = []
            
        manager.metadata = {
            'filename': filename,
            'columns': column_names,
            'rows': len(df) if df is not None else 0,
            'is_preprocessed': True,
            'original_file_type': original_file_type
        }
        
        return manager
    
    @classmethod
    def from_json(cls, json_input: Any, file_id: str, original_filename: str) -> 'SpreadsheetManager':
        """Create SpreadsheetManager from JSON string or dictionary."""
        parsed_json_data: Dict[str, Any]
        if isinstance(json_input, str):
            parsed_json_data = json.loads(json_input)
        elif isinstance(json_input, dict):
            parsed_json_data = json_input
        else:
            raise TypeError("json_input must be a JSON string or a dictionary")

        data_list = parsed_json_data.get('data')
        headers = parsed_json_data.get('headers')

        df: pd.DataFrame
        if data_list is not None:
            if headers is not None:
                df = pd.DataFrame(data_list, columns=headers)
            else:
                df = pd.DataFrame(data_list)
        elif headers is not None:
            df = pd.DataFrame(columns=headers)
        else:
            df = pd.DataFrame()
            
        return cls(file_id, original_filename, df)
    
    # Data Management Methods
    def get_data(self) -> Optional[pd.DataFrame]:
        """Get spreadsheet data as DataFrame."""
        return self.data_df
    
    def set_data(self, data_df: pd.DataFrame) -> None:
        """Update spreadsheet data."""
        self.data_df = data_df
        column_names = [f"Column_{i}" for i in range(len(data_df.columns))]
        self.metadata['columns'] = column_names
        self.metadata['rows'] = len(data_df)
    
    def get_metadata(self) -> Dict[str, Any]:
        """Get spreadsheet metadata."""
        return self.metadata
    
    def to_json(self, save_to_file: bool = False, file_manager = None) -> dict:
        """Convert spreadsheet to JSON format."""
        if self.data_df is not None:
            df_copy = self.data_df.copy()
            
            for col in df_copy.columns:
                if pd.api.types.is_datetime64_any_dtype(df_copy[col]):
                    df_copy[col] = df_copy[col].dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')
            
            df_dict = df_copy.replace({pd.NA: None, float('nan'): None}).to_dict(orient='records')
            headers = df_copy.columns.tolist()
        else:
            df_dict = []
            headers = []
        
        json_data = {
            'file_id': self.file_id,
            'original_filename': self.original_filename,
            'headers': headers,
            'data': df_dict,
            'metadata': self.get_metadata()
        }
        
        if save_to_file and file_manager:
            file_manager.save_json_data(json_data, f"spreadsheet_{self.file_id}")
            
        return json_data
    
    # File Parsing Methods
    def parse_file(self, file_path: str) -> Tuple[pd.DataFrame, List[Dict[str, Any]], str]:
        """Parse spreadsheet file extracting all values as plain text."""
        self.logger.info(f"Starting comprehensive parsing of file: {file_path}")
        
        file_ext = os.path.splitext(file_path)[1].lower()
        original_file_type = file_ext
        
        if file_ext in ['.xlsx', '.xls']:
            df, sheets = self._parse_excel_comprehensive(file_path)
        elif file_ext == '.csv':
            df, sheets = self._parse_csv_comprehensive(file_path)
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")
        
        return df, sheets, original_file_type
    
    def _parse_excel_comprehensive(self, file_path: str) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
        """Comprehensive Excel parsing handling all customizations."""
        try:
            self.logger.info("Loading Excel workbook with comprehensive processing...")
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheets_data = []
            main_df = None
            
            for sheet_name in workbook.sheetnames:
                self.logger.info(f"Processing sheet: {sheet_name}")
                sheet = workbook[sheet_name]
                
                extracted_data = self._extract_sheet_values_comprehensive(sheet)
                
                if extracted_data:
                    df = self._create_clean_dataframe(extracted_data)
                    df = self._clean_excel_data(df)
                    
                    sheet_info = {
                        'name': sheet_name,
                        'data': df
                    }
                    sheets_data.append(sheet_info)
                    
                    if main_df is None and not df.empty:
                        main_df = df
            
            workbook.close()
            
            if main_df is None:
                self.logger.warning("No data found in any sheet, creating empty DataFrame")
                main_df = pd.DataFrame()
            
            self.logger.info(f"Successfully processed {len(sheets_data)} sheets")
            return main_df, sheets_data
            
        except Exception as e:
            self.logger.error(f"Error in comprehensive Excel processing {file_path}: {str(e)}")
            return self._fallback_excel_processing_enhanced(file_path)
    
    def _extract_sheet_values_comprehensive(self, sheet) -> List[List[str]]:
        """Extract all values from worksheet, handling merged cells and Excel customizations."""
        self.logger.debug(f"Extracting values from sheet with max_row={sheet.max_row}, max_col={sheet.max_column}")
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        if max_row == 1 and max_col == 1 and sheet.cell(1, 1).value is None:
            self.logger.debug("Sheet appears to be empty")
            return []
        
        merged_cell_values = {}
        merged_ranges_info = []
        
        for merged_range in sheet.merged_cells.ranges:
            top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
            value = self._cell_value_to_string_enhanced(top_left_cell.value)
            
            merged_ranges_info.append({
                'range': str(merged_range),
                'value': value,
                'size': (merged_range.max_row - merged_range.min_row + 1, 
                        merged_range.max_col - merged_range.min_col + 1)
            })
            
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cell_values[(row, col)] = value
        
        if merged_ranges_info:
            self.logger.debug(f"Found {len(merged_ranges_info)} merged cell ranges")
        
        data = []
        empty_row_count = 0
        
        for row in range(1, max_row + 1):
            row_data = []
            row_has_data = False
            
            for col in range(1, max_col + 1):
                if (row, col) in merged_cell_values:
                    value = merged_cell_values[(row, col)]
                else:
                    cell = sheet.cell(row, col)
                    value = self._cell_value_to_string_enhanced(cell.value)
                
                if value and value.strip():
                    row_has_data = True
                
                row_data.append(value)
            
            if not row_has_data:
                empty_row_count += 1
            else:
                empty_row_count = 0
            
            data.append(row_data)
        
        data = self._clean_extracted_data(data)
        
        self.logger.debug(f"Extracted {len(data)} rows with {len(data[0]) if data else 0} columns")
        return data
    
    def _cell_value_to_string_enhanced(self, value) -> str:
        """Enhanced conversion of cell value to string, handling various data types."""
        if value is None:
            return ''
        elif isinstance(value, str):
            cleaned = value.strip()
            cleaned = re.sub(r'^[\s\u00A0\u2000-\u200F\u2028-\u202F\u205F\u3000]+', '', cleaned)
            cleaned = re.sub(r'[\s\u00A0\u2000-\u200F\u2028-\u202F\u205F\u3000]+$', '', cleaned)
            return cleaned
        elif isinstance(value, (int, float)):
            if isinstance(value, float):
                if value.is_integer():
                    return str(int(value))
                if abs(value) < 1e-6 or abs(value) > 1e15:
                    return f"{value:.2e}"
                else:
                    return f"{value:.10g}"
            return str(value)
        elif hasattr(value, 'strftime'):
            try:
                if hasattr(value, 'time') and value.time() != value.time().replace(hour=0, minute=0, second=0, microsecond=0):
                    return value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    return value.strftime('%Y-%m-%d')
            except:
                return str(value)
        else:
            return str(value).strip()
    
    def _clean_extracted_data(self, data: List[List[str]]) -> List[List[str]]:
        """Clean extracted data by removing empty trailing rows and columns."""
        if not data:
            return data
        
        while data and all(cell == '' for cell in data[-1]):
            data.pop()
        
        if not data:
            return data
        
        while data and all(row and row[-1] == '' for row in data):
            for row in data:
                if row:
                    row.pop()
        
        while data and all(cell == '' for cell in data[0]):
            data.pop(0)
        
        if data:
            while data and all(row and row[0] == '' for row in data):
                for row in data:
                    if row:
                        row.pop(0)
        
        return data
    
    def _create_clean_dataframe(self, data: List[List[str]]) -> pd.DataFrame:
        """Create clean DataFrame from extracted values, ensuring all rows have same length."""
        if not data:
            return pd.DataFrame()
        
        max_cols = max(len(row) for row in data) if data else 0
        
        padded_data = []
        for row in data:
            padded_row = row + [''] * (max_cols - len(row))
            padded_data.append(padded_row)
        
        df = pd.DataFrame(padded_data)
        
        for col in df.columns:
            df[col] = df[col].astype(str)
            df[col] = df[col].replace(['nan', 'NaN', 'NaT', '<NA>'], '')
            df[col] = df[col].str.strip()
        
        return df
    
    def _clean_excel_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply Excel-specific cleaning to the DataFrame."""
        if df.empty:
            return df
        
        df_cleaned = df.copy()
        rows_to_keep = []
        
        for idx, row in df_cleaned.iterrows():
            row_values = [str(val).strip().lower() for val in row.values if str(val).strip()]
            
            if not row_values:
                continue
            
            is_header_row = False
            if len(row_values) == 1:
                for pattern in self.header_patterns:
                    if re.match(pattern, row_values[0], re.IGNORECASE):
                        is_header_row = True
                        break
            
            if not is_header_row:
                rows_to_keep.append(idx)
        
        if rows_to_keep:
            df_cleaned = df_cleaned.iloc[rows_to_keep].reset_index(drop=True)
        
        return df_cleaned
    
    def _parse_csv_comprehensive(self, file_path: str) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
        """Comprehensive CSV parsing with multiple encoding and delimiter attempts."""
        try:
            self.logger.info(f"Starting comprehensive CSV parsing for: {file_path}")
            
            encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1', 'cp850']
            delimiters = [',', ';', '\t', '|', ':', ' ']
            
            df = None
            successful_config = None
            
            for encoding in encodings:
                for delimiter in delimiters:
                    try:
                        test_df = pd.read_csv(
                            file_path,
                            encoding=encoding,
                            delimiter=delimiter,
                            header=None,
                            dtype=str,
                            keep_default_na=False,
                            na_filter=False,
                            skipinitialspace=True,
                            skip_blank_lines=False,
                            quoting=csv.QUOTE_MINIMAL,
                            doublequote=True,
                            escapechar=None
                        )
                        
                        if self._validate_csv_parsing(test_df):
                            df = test_df
                            successful_config = {'encoding': encoding, 'delimiter': delimiter}
                            self.logger.info(f"Successfully parsed CSV with encoding={encoding}, delimiter='{delimiter}'")
                            break
                            
                    except Exception as e:
                        self.logger.debug(f"Failed with encoding={encoding}, delimiter='{delimiter}': {str(e)}")
                        continue
                
                if df is not None:
                    break
            
            if df is None:
                self.logger.warning("Standard parsing failed, attempting fallback parsing")
                df = self._fallback_csv_parsing(file_path)
                
            if df is None or df.empty:
                self.logger.warning("No data found in CSV file")
                df = pd.DataFrame()
            else:
                df = self._clean_csv_data(df)
                self.logger.info(f"Successfully processed CSV: {len(df)} rows, {len(df.columns)} columns")
            
            sheets_data = [{
                'name': 'Sheet1',
                'data': df
            }]
            
            return df, sheets_data
            
        except Exception as e:
            self.logger.error(f"Error in comprehensive CSV parsing {file_path}: {str(e)}")
            return pd.DataFrame(), []
    
    def _validate_csv_parsing(self, df: pd.DataFrame) -> bool:
        """Validate if CSV parsing was successful by checking data characteristics."""
        if df is None or df.empty:
            return False
        
        if len(df.columns) == 1 and len(df) == 1:
            return False
        
        non_empty_cells = 0
        total_cells = len(df) * len(df.columns)
        
        for col in df.columns:
            non_empty_cells += (df[col].astype(str).str.strip() != '').sum()
        
        if total_cells > 0 and (non_empty_cells / total_cells) < 0.1:
            return False
        
        return True
    
    def _fallback_csv_parsing(self, file_path: str) -> pd.DataFrame:
        """Fallback CSV parsing method for problematic files."""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            
            if not lines:
                return pd.DataFrame()
            
            sample_lines = lines[:5]
            potential_delimiters = [',', ';', '\t', '|']
            delimiter_counts = {}
            
            for delimiter in potential_delimiters:
                count = sum(line.count(delimiter) for line in sample_lines)
                delimiter_counts[delimiter] = count
            
            best_delimiter = max(delimiter_counts, key=delimiter_counts.get) if delimiter_counts else ','
            
            data = []
            for line in lines:
                row = [cell.strip().strip('"').strip("'") for cell in line.split(best_delimiter)]
                data.append(row)
            
            if data:
                max_cols = max(len(row) for row in data)
                padded_data = [row + [''] * (max_cols - len(row)) for row in data]
                
                df = pd.DataFrame(padded_data)
                return df
            
        except Exception as e:
            self.logger.error(f"Fallback CSV parsing failed: {str(e)}")
        
        return pd.DataFrame()
    
    def _clean_csv_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Apply comprehensive cleaning to CSV data."""
        if df.empty:
            return df
        
        df_cleaned = df.copy()
        
        for col in df_cleaned.columns:
            df_cleaned[col] = df_cleaned[col].astype(str)
            df_cleaned[col] = df_cleaned[col].str.strip().str.strip('"').str.strip("'").str.strip()
            df_cleaned[col] = df_cleaned[col].replace(['nan', 'NaN', 'NULL', 'null', 'N/A', 'n/a', '#N/A', '#NULL!'], '')
        
        non_empty_rows = ~(df_cleaned == '').all(axis=1)
        if non_empty_rows.any():
            df_cleaned = df_cleaned[non_empty_rows].reset_index(drop=True)
        
        non_empty_cols = ~(df_cleaned == '').all(axis=0)
        if non_empty_cols.any():
            df_cleaned = df_cleaned.loc[:, non_empty_cols]
            df_cleaned.columns = range(len(df_cleaned.columns))
        
        return df_cleaned
    
    def _fallback_excel_processing_enhanced(self, file_path: str) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
        """Enhanced fallback Excel processing using pandas when openpyxl fails."""
        try:
            self.logger.info("Attempting enhanced fallback Excel processing...")
            
            excel_file = pd.ExcelFile(file_path)
            sheets_data = []
            main_df = None
            
            for sheet_name in excel_file.sheet_names:
                self.logger.debug(f"Processing sheet in fallback mode: {sheet_name}")
                
                df = None
                
                try:
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        header=None,
                        dtype=str,
                        keep_default_na=False,
                        na_filter=False
                    )
                except Exception as e:
                    self.logger.debug(f"Approach 1 failed for {sheet_name}: {str(e)}")
                
                if df is None or df.empty:
                    try:
                        df = pd.read_excel(
                            excel_file,
                            sheet_name=sheet_name,
                            header=None,
                            skiprows=1,
                            dtype=str,
                            keep_default_na=False,
                            na_filter=False
                        )
                    except Exception as e:
                        self.logger.debug(f"Approach 2 failed for {sheet_name}: {str(e)}")
                
                if df is None or df.empty:
                    try:
                        df = pd.read_excel(
                            excel_file,
                            sheet_name=sheet_name,
                            dtype=str
                        )
                        if not df.empty:
                            header_row = pd.DataFrame([df.columns.tolist()], columns=df.columns)
                            df = pd.concat([header_row, df], ignore_index=True)
                            df.columns = range(len(df.columns))
                    except Exception as e:
                        self.logger.debug(f"Approach 3 failed for {sheet_name}: {str(e)}")
                
                if df is not None and not df.empty:
                    df = self._clean_fallback_excel_data(df)
                    
                    sheet_info = {
                        'name': sheet_name,
                        'data': df
                    }
                    sheets_data.append(sheet_info)
                    
                    if main_df is None:
                        main_df = df
            
            excel_file.close()
            
            if main_df is None:
                self.logger.warning("Enhanced fallback processing found no data")
                main_df = pd.DataFrame()
            
            return main_df, sheets_data
            
        except Exception as e:
            self.logger.error(f"Enhanced fallback Excel processing failed for {file_path}: {str(e)}")
            return pd.DataFrame(), []
    
    def _clean_fallback_excel_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean data from fallback Excel processing."""
        if df.empty:
            return df
        
        df_cleaned = df.copy()
        
        for col in df_cleaned.columns:
            df_cleaned[col] = df_cleaned[col].astype(str)
            df_cleaned[col] = df_cleaned[col].replace(['nan', 'NaN', 'NaT', '<NA>', 'None'], '')
            df_cleaned[col] = df_cleaned[col].str.strip()
        
        non_empty_rows = ~(df_cleaned == '').all(axis=1)
        if non_empty_rows.any():
            df_cleaned = df_cleaned[non_empty_rows].reset_index(drop=True)
        
        non_empty_cols = ~(df_cleaned == '').all(axis=0)
        if non_empty_cols.any():
            df_cleaned = df_cleaned.loc[:, non_empty_cols]
            df_cleaned.columns = range(len(df_cleaned.columns))
        
        return df_cleaned
    
    # File Saving Methods
    def save(self, save_dir: str, format: str = None) -> str:
        """Save spreadsheet to file in the original format or specified format."""
        if self.data_df is None:
            raise ValueError("Cannot save spreadsheet: no data loaded.")
            
        os.makedirs(save_dir, exist_ok=True)
        
        if format is None:
            if self.original_file_type and self.original_file_type.lower() in ['.xlsx', '.xls']:
                format = 'xlsx'
                file_extension = '.xlsx'
            elif self.original_file_type and self.original_file_type.lower() == '.csv':
                format = 'csv'
                file_extension = '.csv'
            else:
                format = 'csv'
                file_extension = '.csv'
        else:
            file_extension = f'.{format}'
        
        file_path = os.path.join(save_dir, f"{self.file_id}{file_extension}")
        
        if format == 'xlsx':
            self.save_as_original_format(self.data_df, file_path, '.xlsx')
        elif format == 'csv':
            self.save_as_original_format(self.data_df, file_path, '.csv')
        else:
            raise ValueError(f"Unsupported format: {format}")
        
        self.file_path = file_path
        return file_path
    
    def save_as_clean_csv(self, df: pd.DataFrame, output_path: str) -> None:
        """Save DataFrame as clean CSV with enhanced options."""
        try:
            self.logger.info(f"Saving clean CSV to: {output_path}")
            
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            df.to_csv(
                output_path,
                index=False,
                header=False,
                encoding='utf-8',
                quoting=csv.QUOTE_MINIMAL,
                quotechar='"',
                lineterminator='\n',
                escapechar=None,
                doublequote=True
            )
            
            self.logger.info(f"Successfully saved {len(df)} rows and {len(df.columns)} columns to CSV")
            
        except Exception as e:
            self.logger.error(f"Error saving clean CSV to {output_path}: {str(e)}")
            raise
    
    def save_as_original_format(self, df: pd.DataFrame, output_path: str, original_file_type: str) -> None:
        """Save DataFrame back to the original file format."""
        try:
            self.logger.info(f"Saving file as original format: {original_file_type}")
            
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            if original_file_type.lower() in ['.xlsx', '.xls']:
                excel_output_path = os.path.splitext(output_path)[0] + '.xlsx'
                
                with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
                    df.to_excel(
                        writer,
                        index=False,
                        header=False,
                        sheet_name='Sheet1'
                    )
                
                self.logger.info(f"Successfully saved {len(df)} rows and {len(df.columns)} columns to Excel format")
                
            elif original_file_type.lower() == '.csv':
                csv_output_path = os.path.splitext(output_path)[0] + '.csv'
                self.save_as_clean_csv(df, csv_output_path)
                
            else:
                self.logger.warning(f"Unsupported original format {original_file_type}, defaulting to CSV")
                csv_output_path = os.path.splitext(output_path)[0] + '.csv'
                self.save_as_clean_csv(df, csv_output_path)
            
        except Exception as e:
            self.logger.error(f"Error saving file as original format {original_file_type}: {str(e)}")
            raise
    
    # Utility Methods
    def get_parsing_summary(self, df: pd.DataFrame, sheets: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Generate a summary of the parsing results."""
        summary = {
            'total_sheets': len(sheets),
            'main_sheet_rows': len(df) if df is not None else 0,
            'main_sheet_cols': len(df.columns) if df is not None and not df.empty else 0,
            'total_cells': 0,
            'non_empty_cells': 0,
            'sheets_info': []
        }
        
        for sheet_info in sheets:
            sheet_df = sheet_info['data']
            if sheet_df is not None and not sheet_df.empty:
                rows = len(sheet_df)
                cols = len(sheet_df.columns)
                total_cells = rows * cols
                non_empty = 0
                
                for col in sheet_df.columns:
                    non_empty += (sheet_df[col].astype(str).str.strip() != '').sum()
                
                sheet_summary = {
                    'name': sheet_info['name'],
                    'rows': rows,
                    'columns': cols,
                    'total_cells': total_cells,
                    'non_empty_cells': non_empty,
                    'data_density': (non_empty / total_cells * 100) if total_cells > 0 else 0
                }
                
                summary['sheets_info'].append(sheet_summary)
                summary['total_cells'] += total_cells
                summary['non_empty_cells'] += non_empty
        
        summary['overall_data_density'] = (
            summary['non_empty_cells'] / summary['total_cells'] * 100 
            if summary['total_cells'] > 0 else 0
        )
        
        return summary
    
    @staticmethod
    def detect_file_format(file_path: str) -> str:
        """Detect the format of a spreadsheet file."""
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext in ['.xlsx', '.xls']:
            return 'excel'
        elif file_ext == '.csv':
            return 'csv'
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")
    
    @staticmethod
    def parse_from_pandas(df: pd.DataFrame) -> str:
        """Convert a pandas DataFrame to JSON string."""
        df_clean = df.copy()
        for col in df_clean.columns:
            df_clean[col] = df_clean[col].astype(str).fillna('')
        
        return df_clean.to_json(orient='records')
    
    # Legacy methods for backward compatibility
    @staticmethod
    def parse_to_json(file_path: str) -> Tuple[str, Dict[str, Any]]:
        """Parse spreadsheet file to JSON (legacy method for backward compatibility)."""
        manager = SpreadsheetManager()
        df, sheets, original_file_type = manager.parse_file(file_path)
        
        if df.empty:
            return '[]', {'columns': [], 'rows': 0, 'file_type': original_file_type[1:]}
        
        data = df.to_json(orient='records')
        
        metadata = {
            'columns': [f"Column_{i}" for i in range(len(df.columns))],
            'rows': len(df),
            'file_type': original_file_type[1:]
        }
        
        return data, metadata


# Backward compatibility aliases
class SpreadsheetParser(SpreadsheetManager):
    """Backward compatibility alias for SpreadsheetParser."""
    pass

class Spreadsheet(SpreadsheetManager):
    """Backward compatibility alias for Spreadsheet."""
    pass
