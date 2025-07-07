"""
Spreadsheet Parser module
-----------------------
Handles comprehensive parsing and preprocessing of spreadsheet files to extract plain text data
and convert everything to clean CSV format, removing all formatting issues, 
merged cells, headers, group headers, and other customizations that can cause parsing problems.
"""

import os
import pandas as pd
import openpyxl
import csv
import io
import re
from typing import List, Dict, Any, Optional, Union, Tuple
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import logging

class SpreadsheetParser:
    """
    Comprehensive parser for spreadsheet files that handles all types of customizations
    including merged cells, merged rows/columns, group headers, formatting, and other issues
    that can cause parsing problems. Extracts all values as plain text and converts to clean CSV.
    """
    
    def __init__(self):
        """Initialize the parser"""
        self.logger = logging.getLogger(__name__)
        # Pattern to detect potential header-like content
        self.header_patterns = [
            r'^(total|sum|subtotal|grand total|net|gross)$',
            r'^(header|title|label)$',
            r'^\s*$',  # Empty cells
            r'^[A-Z\s]+$',  # All caps (potential headers)
        ]
    
    def parse_file(self, file_path: str) -> Tuple[pd.DataFrame, List[Dict[str, Any]], str]:
        """
        Comprehensive parsing of a spreadsheet file to extract all values as plain text.
        Handles all types of spreadsheet customizations that can cause parsing issues.
        
        Args:
            file_path: Path to the spreadsheet file
            
        Returns:
            Tuple[pd.DataFrame, List[Dict], str]: Main dataframe, list of all sheets, and original file type
        """
        self.logger.info(f"Starting comprehensive parsing of file: {file_path}")
        
        file_ext = os.path.splitext(file_path)[1].lower()
        original_file_type = file_ext  # Store the original file type
        
        if file_ext in ['.xlsx', '.xls']:
            df, sheets = self._parse_excel_comprehensive(file_path)
        elif file_ext == '.csv':
            df, sheets = self._parse_csv_comprehensive(file_path)
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")
        
        return df, sheets, original_file_type
    
    def _parse_excel_comprehensive(self, file_path: str) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
        """
        Comprehensive Excel parsing that handles all types of customizations
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple[pd.DataFrame, List[Dict]]: Main dataframe and list of all sheets
        """
        try:
            self.logger.info("Loading Excel workbook with comprehensive processing...")
            
            # Load workbook with data_only=True to get calculated values
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            sheets_data = []
            main_df = None
            
            for sheet_name in workbook.sheetnames:
                self.logger.info(f"Processing sheet: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # Extract all cell values with comprehensive handling
                extracted_data = self._extract_sheet_values_comprehensive(sheet)
                
                if extracted_data:
                    # Create clean DataFrame from extracted data
                    df = self._create_clean_dataframe(extracted_data)
                    
                    # Additional cleaning for Excel-specific issues
                    df = self._clean_excel_data(df)
                    
                    # Store sheet info
                    sheet_info = {
                        'name': sheet_name,
                        'data': df
                    }
                    sheets_data.append(sheet_info)
                    
                    # Use first non-empty sheet as main dataframe
                    if main_df is None and not df.empty:
                        main_df = df
            
            workbook.close()
            
            # If no data found, create empty dataframe
            if main_df is None:
                self.logger.warning("No data found in any sheet, creating empty DataFrame")
                main_df = pd.DataFrame()
            
            self.logger.info(f"Successfully processed {len(sheets_data)} sheets")
            return main_df, sheets_data
            
        except Exception as e:
            self.logger.error(f"Error in comprehensive Excel processing {file_path}: {str(e)}")
            # Fallback to enhanced processing
            return self._fallback_excel_processing_enhanced(file_path)
    
    def _extract_sheet_values_comprehensive(self, sheet) -> List[List[str]]:
        """
        Comprehensive extraction of all values from a worksheet, handling merged cells,
        merged rows/columns, group headers, and other Excel customizations
        
        Args:
            sheet: openpyxl worksheet object
            
        Returns:
            List[List[str]]: 2D list of cell values as strings
        """
        self.logger.debug(f"Extracting values from sheet with max_row={sheet.max_row}, max_col={sheet.max_column}")
        
        # Find the actual data range (skip completely empty sheets)
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # Handle case where sheet appears empty
        if max_row == 1 and max_col == 1 and sheet.cell(1, 1).value is None:
            self.logger.debug("Sheet appears to be empty")
            return []
        
        # Create a comprehensive mapping of merged cell ranges to their values
        merged_cell_values = {}
        merged_ranges_info = []
        
        # Process all merged cells
        for merged_range in sheet.merged_cells.ranges:
            # Get the top-left cell value
            top_left_cell = sheet.cell(merged_range.min_row, merged_range.min_col)
            value = self._cell_value_to_string_enhanced(top_left_cell.value)
            
            # Store merged range info for debugging
            merged_ranges_info.append({
                'range': str(merged_range),
                'value': value,
                'size': (merged_range.max_row - merged_range.min_row + 1, 
                        merged_range.max_col - merged_range.min_col + 1)
            })
            
            # Map all cells in the merged range to this value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cell_values[(row, col)] = value
        
        if merged_ranges_info:
            self.logger.debug(f"Found {len(merged_ranges_info)} merged cell ranges")
        
        # Extract all cell values with enhanced processing
        data = []
        empty_row_count = 0
        
        for row in range(1, max_row + 1):
            row_data = []
            row_has_data = False
            
            for col in range(1, max_col + 1):
                if (row, col) in merged_cell_values:
                    # Use merged cell value
                    value = merged_cell_values[(row, col)]
                else:
                    # Get regular cell value with enhanced processing
                    cell = sheet.cell(row, col)
                    value = self._cell_value_to_string_enhanced(cell.value)
                
                # Check if this cell contains actual data
                if value and value.strip():
                    row_has_data = True
                
                row_data.append(value)
            
            # Track empty rows but include them in case they separate data sections
            if not row_has_data:
                empty_row_count += 1
            else:
                empty_row_count = 0
            
            data.append(row_data)
        
        # Post-process the extracted data
        data = self._clean_extracted_data(data)
        
        self.logger.debug(f"Extracted {len(data)} rows with {len(data[0]) if data else 0} columns")
        return data
    
    def _cell_value_to_string_enhanced(self, value) -> str:
        """
        Enhanced conversion of cell value to string, handling various data types
        and Excel-specific formatting issues
        
        Args:
            value: Cell value of any type
            
        Returns:
            str: Clean string representation of the value
        """
        if value is None:
            return ''
        elif isinstance(value, str):
            # Clean up string values
            cleaned = value.strip()
            # Remove common Excel artifacts
            cleaned = re.sub(r'^[\s\u00A0\u2000-\u200F\u2028-\u202F\u205F\u3000]+', '', cleaned)
            cleaned = re.sub(r'[\s\u00A0\u2000-\u200F\u2028-\u202F\u205F\u3000]+$', '', cleaned)
            return cleaned
        elif isinstance(value, (int, float)):
            # Handle numbers with enhanced precision
            if isinstance(value, float):
                if value.is_integer():
                    return str(int(value))
                # Format floats to avoid scientific notation for reasonable ranges
                if abs(value) < 1e-6 or abs(value) > 1e15:
                    return f"{value:.2e}"
                else:
                    return f"{value:.10g}"
            return str(value)
        elif hasattr(value, 'strftime'):
            # Handle datetime objects with better formatting
            try:
                if hasattr(value, 'time') and value.time() != value.time().replace(hour=0, minute=0, second=0, microsecond=0):
                    return value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    return value.strftime('%Y-%m-%d')
            except:
                return str(value)
        else:
            # Handle other types (boolean, etc.)
            return str(value).strip()
    
    def _clean_extracted_data(self, data: List[List[str]]) -> List[List[str]]:
        """
        Clean the extracted data by removing completely empty trailing rows and columns,
        and handling common spreadsheet artifacts
        
        Args:
            data: 2D list of extracted values
            
        Returns:
            List[List[str]]: Cleaned data
        """
        if not data:
            return data
        
        # Remove completely empty rows from the end
        while data and all(cell == '' for cell in data[-1]):
            data.pop()
        
        if not data:
            return data
        
        # Remove completely empty columns from the end
        while data and all(row and row[-1] == '' for row in data):
            for row in data:
                if row:
                    row.pop()
        
        # Remove leading empty rows (common in Excel with headers)
        while data and all(cell == '' for cell in data[0]):
            data.pop(0)
        
        # Remove leading empty columns
        if data:
            while data and all(row and row[0] == '' for row in data):
                for row in data:
                    if row:
                        row.pop(0)
        
        return data
    
    def _create_clean_dataframe(self, data: List[List[str]]) -> pd.DataFrame:
        """
        Create a clean DataFrame from extracted values, ensuring all rows have same length
        and applying additional data cleaning
        
        Args:
            data: 2D list of string values
            
        Returns:
            pd.DataFrame: Clean DataFrame with enhanced processing
        """
        if not data:
            return pd.DataFrame()
        
        # Find the maximum row length
        max_cols = max(len(row) for row in data) if data else 0
        
        # Pad all rows to have the same length
        padded_data = []
        for row in data:
            padded_row = row + [''] * (max_cols - len(row))
            padded_data.append(padded_row)
        
        # Create DataFrame without headers (all data treated as values)
        df = pd.DataFrame(padded_data)
        
        # Enhanced data cleaning
        for col in df.columns:
            # Convert to string and clean
            df[col] = df[col].astype(str)
            # Replace pandas NaN representations
            df[col] = df[col].replace(['nan', 'NaN', 'NaT', '<NA>'], '')
            # Clean whitespace
            df[col] = df[col].str.strip()
        
        return df
    
    def _clean_excel_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply Excel-specific cleaning to the DataFrame
        
        Args:
            df: DataFrame to clean
            
        Returns:
            pd.DataFrame: Cleaned DataFrame
        """
        if df.empty:
            return df
        
        df_cleaned = df.copy()
        
        # Remove rows that are likely headers or formatting artifacts
        rows_to_keep = []
        
        for idx, row in df_cleaned.iterrows():
            row_values = [str(val).strip().lower() for val in row.values if str(val).strip()]
            
            # Skip if row is completely empty
            if not row_values:
                continue
            
            # Skip if row contains only header-like patterns
            is_header_row = False
            if len(row_values) == 1:
                for pattern in self.header_patterns:
                    if re.match(pattern, row_values[0], re.IGNORECASE):
                        is_header_row = True
                        break
            
            if not is_header_row:
                rows_to_keep.append(idx)
        
        if rows_to_keep:
            df_cleaned = df_cleaned.iloc[rows_to_keep].reset_index(drop=True)
        
        return df_cleaned
    
    def _parse_csv_comprehensive(self, file_path: str) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
        """
        Comprehensive CSV parsing to ensure clean format and handle various CSV issues
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            Tuple[pd.DataFrame, List[Dict]]: Main dataframe and list with single sheet
        """
        try:
            self.logger.info(f"Starting comprehensive CSV parsing for: {file_path}")
            
            # Try different encodings and delimiters systematically
            encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1', 'cp850']
            delimiters = [',', ';', '\t', '|', ':', ' ']
            
            df = None
            successful_config = None
            
            for encoding in encodings:
                for delimiter in delimiters:
                    try:
                        # Read CSV without headers to treat everything as data
                        test_df = pd.read_csv(
                            file_path,
                            encoding=encoding,
                            delimiter=delimiter,
                            header=None,  # No headers
                            dtype=str,    # Everything as string
                            keep_default_na=False,  # Don't convert to NaN
                            na_filter=False,  # Don't filter NA values
                            skipinitialspace=True,  # Skip whitespace after delimiter
                            skip_blank_lines=False,  # Keep blank lines for structure
                            quoting=csv.QUOTE_MINIMAL,
                            doublequote=True,
                            escapechar=None
                        )
                        
                        # Validate if this looks like valid data
                        if self._validate_csv_parsing(test_df):
                            df = test_df
                            successful_config = {'encoding': encoding, 'delimiter': delimiter}
                            self.logger.info(f"Successfully parsed CSV with encoding={encoding}, delimiter='{delimiter}'")
                            break
                            
                    except Exception as e:
                        self.logger.debug(f"Failed with encoding={encoding}, delimiter='{delimiter}': {str(e)}")
                        continue
                
                if df is not None:
                    break
            
            # Fallback parsing if no configuration worked
            if df is None:
                self.logger.warning("Standard parsing failed, attempting fallback parsing")
                df = self._fallback_csv_parsing(file_path)
                
            if df is None or df.empty:
                self.logger.warning("No data found in CSV file")
                df = pd.DataFrame()
            else:
                # Apply comprehensive cleaning to the successfully parsed data
                df = self._clean_csv_data(df)
                
                self.logger.info(f"Successfully processed CSV: {len(df)} rows, {len(df.columns)} columns")
            
            # Create sheets list with single sheet
            sheets_data = [{
                'name': 'Sheet1',
                'data': df
            }]
            
            return df, sheets_data
            
        except Exception as e:
            self.logger.error(f"Error in comprehensive CSV parsing {file_path}: {str(e)}")
            # Create empty dataframe as final fallback
            return pd.DataFrame(), []
    
    def _validate_csv_parsing(self, df: pd.DataFrame) -> bool:
        """
        Validate if CSV parsing was successful by checking data characteristics
        
        Args:
            df: DataFrame to validate
            
        Returns:
            bool: True if parsing appears successful
        """
        if df is None or df.empty:
            return False
        
        # Check for reasonable dimensions
        if len(df.columns) == 1 and len(df) == 1:
            return False
        
        # Check if we have reasonable data distribution
        non_empty_cells = 0
        total_cells = len(df) * len(df.columns)
        
        for col in df.columns:
            non_empty_cells += (df[col].astype(str).str.strip() != '').sum()
        
        # At least 10% of cells should have data
        if total_cells > 0 and (non_empty_cells / total_cells) < 0.1:
            return False
        
        return True
    
    def _fallback_csv_parsing(self, file_path: str) -> pd.DataFrame:
        """
        Fallback CSV parsing method for problematic files
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            pd.DataFrame: Parsed dataframe or empty if failed
        """
        try:
            # Try reading as raw text and parsing manually
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            
            if not lines:
                return pd.DataFrame()
            
            # Try to detect delimiter from first few lines
            sample_lines = lines[:5]
            potential_delimiters = [',', ';', '\t', '|']
            delimiter_counts = {}
            
            for delimiter in potential_delimiters:
                count = sum(line.count(delimiter) for line in sample_lines)
                delimiter_counts[delimiter] = count
            
            # Use the delimiter that appears most frequently
            best_delimiter = max(delimiter_counts, key=delimiter_counts.get) if delimiter_counts else ','
            
            # Parse manually
            data = []
            for line in lines:
                row = [cell.strip().strip('"').strip("'") for cell in line.split(best_delimiter)]
                data.append(row)
            
            if data:
                # Find max columns
                max_cols = max(len(row) for row in data)
                # Pad rows to same length
                padded_data = [row + [''] * (max_cols - len(row)) for row in data]
                
                df = pd.DataFrame(padded_data)
                return df
            
        except Exception as e:
            self.logger.error(f"Fallback CSV parsing failed: {str(e)}")
        
        return pd.DataFrame()
    
    def _clean_csv_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply comprehensive cleaning to CSV data
        
        Args:
            df: DataFrame to clean
            
        Returns:
            pd.DataFrame: Cleaned DataFrame
        """
        if df.empty:
            return df
        
        df_cleaned = df.copy()
        
        # Clean all string values
        for col in df_cleaned.columns:
            df_cleaned[col] = df_cleaned[col].astype(str)
            # Remove quotes and clean whitespace
            df_cleaned[col] = df_cleaned[col].str.strip().str.strip('"').str.strip("'").str.strip()
            # Replace various representations of missing data
            df_cleaned[col] = df_cleaned[col].replace(['nan', 'NaN', 'NULL', 'null', 'N/A', 'n/a', '#N/A', '#NULL!'], '')
        
        # Remove completely empty rows and columns
        # Remove empty rows
        non_empty_rows = ~(df_cleaned == '').all(axis=1)
        if non_empty_rows.any():
            df_cleaned = df_cleaned[non_empty_rows].reset_index(drop=True)
        
        # Remove empty columns
        non_empty_cols = ~(df_cleaned == '').all(axis=0)
        if non_empty_cols.any():
            df_cleaned = df_cleaned.loc[:, non_empty_cols]
            # Reset column names to be sequential
            df_cleaned.columns = range(len(df_cleaned.columns))
        
        return df_cleaned
    
    def _fallback_excel_processing_enhanced(self, file_path: str) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
        """
        Enhanced fallback Excel processing using pandas when openpyxl fails
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple[pd.DataFrame, List[Dict]]: Main dataframe and list of sheets
        """
        try:
            self.logger.info("Attempting enhanced fallback Excel processing...")
            
            # Read all sheets with multiple approaches
            excel_file = pd.ExcelFile(file_path)
            sheets_data = []
            main_df = None
            
            for sheet_name in excel_file.sheet_names:
                self.logger.debug(f"Processing sheet in fallback mode: {sheet_name}")
                
                # Try multiple reading approaches
                df = None
                
                # Approach 1: Read without headers, everything as string
                try:
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        header=None,
                        dtype=str,
                        keep_default_na=False,
                        na_filter=False
                    )
                except Exception as e:
                    self.logger.debug(f"Approach 1 failed for {sheet_name}: {str(e)}")
                
                # Approach 2: Skip potential header rows
                if df is None or df.empty:
                    try:
                        df = pd.read_excel(
                            excel_file,
                            sheet_name=sheet_name,
                            header=None,
                            skiprows=1,  # Skip first row in case it's problematic
                            dtype=str,
                            keep_default_na=False,
                            na_filter=False
                        )
                    except Exception as e:
                        self.logger.debug(f"Approach 2 failed for {sheet_name}: {str(e)}")
                
                # Approach 3: Use default pandas behavior as last resort
                if df is None or df.empty:
                    try:
                        df = pd.read_excel(
                            excel_file,
                            sheet_name=sheet_name,
                            dtype=str
                        )
                        # Reset index to treat header as data
                        if not df.empty:
                            # Insert the original column names as the first row
                            header_row = pd.DataFrame([df.columns.tolist()], columns=df.columns)
                            df = pd.concat([header_row, df], ignore_index=True)
                            # Reset column names to numeric
                            df.columns = range(len(df.columns))
                    except Exception as e:
                        self.logger.debug(f"Approach 3 failed for {sheet_name}: {str(e)}")
                
                if df is not None and not df.empty:
                    # Apply comprehensive cleaning
                    df = self._clean_fallback_excel_data(df)
                    
                    sheet_info = {
                        'name': sheet_name,
                        'data': df
                    }
                    sheets_data.append(sheet_info)
                    
                    if main_df is None:
                        main_df = df
            
            excel_file.close()
            
            if main_df is None:
                self.logger.warning("Enhanced fallback processing found no data")
                main_df = pd.DataFrame()
            
            return main_df, sheets_data
            
        except Exception as e:
            self.logger.error(f"Enhanced fallback Excel processing failed for {file_path}: {str(e)}")
            return pd.DataFrame(), []
    
    def _clean_fallback_excel_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean data from fallback Excel processing
        
        Args:
            df: DataFrame to clean
            
        Returns:
            pd.DataFrame: Cleaned DataFrame
        """
        if df.empty:
            return df
        
        df_cleaned = df.copy()
        
        # Convert all to string and clean
        for col in df_cleaned.columns:
            df_cleaned[col] = df_cleaned[col].astype(str)
            # Clean various NaN representations
            df_cleaned[col] = df_cleaned[col].replace(['nan', 'NaN', 'NaT', '<NA>', 'None'], '')
            # Clean whitespace
            df_cleaned[col] = df_cleaned[col].str.strip()
        
        # Remove completely empty rows and columns
        # Remove rows where all values are empty
        non_empty_rows = ~(df_cleaned == '').all(axis=1)
        if non_empty_rows.any():
            df_cleaned = df_cleaned[non_empty_rows].reset_index(drop=True)
        
        # Remove columns where all values are empty
        non_empty_cols = ~(df_cleaned == '').all(axis=0)
        if non_empty_cols.any():
            df_cleaned = df_cleaned.loc[:, non_empty_cols]
            # Reset column names
            df_cleaned.columns = range(len(df_cleaned.columns))
        
        return df_cleaned
    
    def save_as_clean_csv(self, df: pd.DataFrame, output_path: str) -> None:
        """
        Save DataFrame as clean CSV with enhanced options
        
        Args:
            df: DataFrame to save
            output_path: Path where to save the CSV
        """
        try:
            self.logger.info(f"Saving clean CSV to: {output_path}")
            
            # Ensure the output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save with enhanced CSV settings
            df.to_csv(
                output_path,
                index=False,
                header=False,  # No headers since we treat everything as data
                encoding='utf-8',
                quoting=csv.QUOTE_MINIMAL,
                quotechar='"',
                lineterminator='\n',
                escapechar=None,
                doublequote=True
            )
            
            self.logger.info(f"Successfully saved {len(df)} rows and {len(df.columns)} columns to CSV")
            
        except Exception as e:
            self.logger.error(f"Error saving clean CSV to {output_path}: {str(e)}")
            raise
    
    def save_as_original_format(self, df: pd.DataFrame, output_path: str, original_file_type: str) -> None:
        """
        Save DataFrame back to the original file format
        
        Args:
            df: DataFrame to save
            output_path: Path where to save the file
            original_file_type: Original file extension (.xlsx, .xls, .csv)
        """
        try:
            self.logger.info(f"Saving file as original format: {original_file_type}")
            
            # Ensure the output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            if original_file_type.lower() in ['.xlsx', '.xls']:
                # Save as Excel file
                # Use .xlsx format regardless of original (.xls or .xlsx)
                excel_output_path = os.path.splitext(output_path)[0] + '.xlsx'
                
                with pd.ExcelWriter(excel_output_path, engine='openpyxl') as writer:
                    df.to_excel(
                        writer,
                        index=False,
                        header=False,  # No headers since we treat everything as data
                        sheet_name='Sheet1'
                    )
                
                self.logger.info(f"Successfully saved {len(df)} rows and {len(df.columns)} columns to Excel format")
                
            elif original_file_type.lower() == '.csv':
                # Save as CSV file
                csv_output_path = os.path.splitext(output_path)[0] + '.csv'
                self.save_as_clean_csv(df, csv_output_path)
                
            else:
                # Default to CSV for unsupported formats
                self.logger.warning(f"Unsupported original format {original_file_type}, defaulting to CSV")
                csv_output_path = os.path.splitext(output_path)[0] + '.csv'
                self.save_as_clean_csv(df, csv_output_path)
            
        except Exception as e:
            self.logger.error(f"Error saving file as original format {original_file_type}: {str(e)}")
            raise
    
    def get_parsing_summary(self, df: pd.DataFrame, sheets: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Generate a summary of the parsing results
        
        Args:
            df: Main dataframe
            sheets: List of sheet information
            
        Returns:
            Dict: Summary information
        """
        summary = {
            'total_sheets': len(sheets),
            'main_sheet_rows': len(df) if df is not None else 0,
            'main_sheet_cols': len(df.columns) if df is not None and not df.empty else 0,
            'total_cells': 0,
            'non_empty_cells': 0,
            'sheets_info': []
        }
        
        for sheet_info in sheets:
            sheet_df = sheet_info['data']
            if sheet_df is not None and not sheet_df.empty:
                rows = len(sheet_df)
                cols = len(sheet_df.columns)
                total_cells = rows * cols
                non_empty = 0
                
                for col in sheet_df.columns:
                    non_empty += (sheet_df[col].astype(str).str.strip() != '').sum()
                
                sheet_summary = {
                    'name': sheet_info['name'],
                    'rows': rows,
                    'columns': cols,
                    'total_cells': total_cells,
                    'non_empty_cells': non_empty,
                    'data_density': (non_empty / total_cells * 100) if total_cells > 0 else 0
                }
                
                summary['sheets_info'].append(sheet_summary)
                summary['total_cells'] += total_cells
                summary['non_empty_cells'] += non_empty
        
        summary['overall_data_density'] = (
            summary['non_empty_cells'] / summary['total_cells'] * 100 
            if summary['total_cells'] > 0 else 0
        )
        
        return summary

    # Legacy methods for backward compatibility
    @staticmethod
    def parse_to_json(file_path: str) -> Tuple[str, Dict[str, Any]]:
        """
        Parse a spreadsheet file to JSON (legacy method for backward compatibility)
        Note: This method now uses the enhanced parser but maintains the old interface

        Args:
            file_path: Path to the spreadsheet file

        Returns:
            Tuple[str, Dict]: JSON representation and metadata
        """
        parser = SpreadsheetParser()
        df, sheets, original_file_type = parser.parse_file(file_path)
        
        if df.empty:
            return '[]', {'columns': [], 'rows': 0, 'file_type': original_file_type[1:]}
        
        # Convert to JSON
        data = df.to_json(orient='records')
        
        # Prepare metadata (no column names since we treat everything as data)
        metadata = {
            'columns': [f"Column_{i}" for i in range(len(df.columns))],
            'rows': len(df),
            'file_type': original_file_type[1:]  # Remove the dot
        }
        
        return data, metadata
    
    @staticmethod
    def detect_file_format(file_path: str) -> str:
        """
        Detect the format of a spreadsheet file

        Args:
            file_path: Path to the spreadsheet file

        Returns:
            str: Detected format (excel, csv)
        """
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext in ['.xlsx', '.xls']:
            return 'excel'
        elif file_ext == '.csv':
            return 'csv'
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")
    
    @staticmethod
    def parse_from_pandas(df: pd.DataFrame) -> str:
        """
        Convert a pandas DataFrame to JSON string

        Args:
            df: Pandas DataFrame

        Returns:
            str: JSON representation of the DataFrame
        """
        # Ensure all data is string type and clean
        df_clean = df.copy()
        for col in df_clean.columns:
            df_clean[col] = df_clean[col].astype(str).fillna('')
        
        return df_clean.to_json(orient='records')
    
    def _cell_value_to_string_enhanced(self, value) -> str:
        """
        Enhanced conversion of cell value to string, handling various data types
        and Excel-specific formatting issues
        
        Args:
            value: Cell value of any type
            
        Returns:
            str: Clean string representation of the value
        """
        if value is None:
            return ''
        elif isinstance(value, str):
            # Clean up string values
            cleaned = value.strip()
            # Remove common Excel artifacts
            cleaned = re.sub(r'^[\s\u00A0\u2000-\u200F\u2028-\u202F\u205F\u3000]+', '', cleaned)
            cleaned = re.sub(r'[\s\u00A0\u2000-\u200F\u2028-\u202F\u205F\u3000]+$', '', cleaned)
            return cleaned
        elif isinstance(value, (int, float)):
            # Handle numbers with enhanced precision
            if isinstance(value, float):
                if value.is_integer():
                    return str(int(value))
                # Format floats to avoid scientific notation for reasonable ranges
                if abs(value) < 1e-6 or abs(value) > 1e15:
                    return f"{value:.2e}"
                else:
                    return f"{value:.10g}"
            return str(value)
        elif hasattr(value, 'strftime'):
            # Handle datetime objects with better formatting
            try:
                if hasattr(value, 'time') and value.time() != value.time().replace(hour=0, minute=0, second=0, microsecond=0):
                    return value.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    return value.strftime('%Y-%m-%d')
            except:
                return str(value)
        else:
            # Handle other types (boolean, etc.)
            return str(value).strip()
    
    def _clean_extracted_data(self, data: List[List[str]]) -> List[List[str]]:
        """
        Clean the extracted data by removing completely empty trailing rows and columns,
        and handling common spreadsheet artifacts
        
        Args:
            data: 2D list of extracted values
            
        Returns:
            List[List[str]]: Cleaned data
        """
        if not data:
            return data
        
        # Remove completely empty rows from the end
        while data and all(cell == '' for cell in data[-1]):
            data.pop()
        
        if not data:
            return data
        
        # Remove completely empty columns from the end
        while data and all(row and row[-1] == '' for row in data):
            for row in data:
                if row:
                    row.pop()
        
        # Remove leading empty rows (common in Excel with headers)
        while data and all(cell == '' for cell in data[0]):
            data.pop(0)
        
        # Remove leading empty columns
        if data:
            while data and all(row and row[0] == '' for row in data):
                for row in data:
                    if row:
                        row.pop(0)
        
        return data
    
    def _create_clean_dataframe(self, data: List[List[str]]) -> pd.DataFrame:
        """
        Create a clean DataFrame from extracted values, ensuring all rows have same length
        and applying additional data cleaning
        
        Args:
            data: 2D list of string values
            
        Returns:
            pd.DataFrame: Clean DataFrame with enhanced processing
        """
        if not data:
            return pd.DataFrame()
        
        # Find the maximum row length
        max_cols = max(len(row) for row in data) if data else 0
        
        # Pad all rows to have the same length
        padded_data = []
        for row in data:
            padded_row = row + [''] * (max_cols - len(row))
            padded_data.append(padded_row)
        
        # Create DataFrame without headers (all data treated as values)
        df = pd.DataFrame(padded_data)
        
        # Enhanced data cleaning
        for col in df.columns:
            # Convert to string and clean
            df[col] = df[col].astype(str)
            # Replace pandas NaN representations
            df[col] = df[col].replace(['nan', 'NaN', 'NaT', '<NA>'], '')
            # Clean whitespace
            df[col] = df[col].str.strip()
        
        return df
    
    def _clean_excel_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply Excel-specific cleaning to the DataFrame
        
        Args:
            df: DataFrame to clean
            
        Returns:
            pd.DataFrame: Cleaned DataFrame
        """
        if df.empty:
            return df
        
        df_cleaned = df.copy()
        
        # Remove rows that are likely headers or formatting artifacts
        rows_to_keep = []
        
        for idx, row in df_cleaned.iterrows():
            row_values = [str(val).strip().lower() for val in row.values if str(val).strip()]
            
            # Skip if row is completely empty
            if not row_values:
                continue
            
            # Skip if row contains only header-like patterns
            is_header_row = False
            if len(row_values) == 1:
                for pattern in self.header_patterns:
                    if re.match(pattern, row_values[0], re.IGNORECASE):
                        is_header_row = True
                        break
            
            if not is_header_row:
                rows_to_keep.append(idx)
        
        if rows_to_keep:
            df_cleaned = df_cleaned.iloc[rows_to_keep].reset_index(drop=True)
        
        return df_cleaned
    
    def _parse_csv_comprehensive(self, file_path: str) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
        """
        Comprehensive CSV parsing to ensure clean format and handle various CSV issues
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            Tuple[pd.DataFrame, List[Dict]]: Main dataframe and list with single sheet
        """
        try:
            self.logger.info(f"Starting comprehensive CSV parsing for: {file_path}")
            
            # Try different encodings and delimiters systematically
            encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1', 'cp850']
            delimiters = [',', ';', '\t', '|', ':', ' ']
            
            df = None
            successful_config = None
            
            for encoding in encodings:
                for delimiter in delimiters:
                    try:
                        # Read CSV without headers to treat everything as data
                        test_df = pd.read_csv(
                            file_path,
                            encoding=encoding,
                            delimiter=delimiter,
                            header=None,  # No headers
                            dtype=str,    # Everything as string
                            keep_default_na=False,  # Don't convert to NaN
                            na_filter=False,  # Don't filter NA values
                            skipinitialspace=True,  # Skip whitespace after delimiter
                            skip_blank_lines=False,  # Keep blank lines for structure
                            quoting=csv.QUOTE_MINIMAL,
                            doublequote=True,
                            escapechar=None
                        )
                        
                        # Validate if this looks like valid data
                        if self._validate_csv_parsing(test_df):
                            df = test_df
                            successful_config = {'encoding': encoding, 'delimiter': delimiter}
                            self.logger.info(f"Successfully parsed CSV with encoding={encoding}, delimiter='{delimiter}'")
                            break
                            
                    except Exception as e:
                        self.logger.debug(f"Failed with encoding={encoding}, delimiter='{delimiter}': {str(e)}")
                        continue
                
                if df is not None:
                    break
            
            # Fallback parsing if no configuration worked
            if df is None:
                self.logger.warning("Standard parsing failed, attempting fallback parsing")
                df = self._fallback_csv_parsing(file_path)
                
            if df is None or df.empty:
                self.logger.warning("No data found in CSV file")
                df = pd.DataFrame()
            else:
                # Apply comprehensive cleaning to the successfully parsed data
                df = self._clean_csv_data(df)
                
                self.logger.info(f"Successfully processed CSV: {len(df)} rows, {len(df.columns)} columns")
            
            # Create sheets list with single sheet
            sheets_data = [{
                'name': 'Sheet1',
                'data': df
            }]
            
            return df, sheets_data
            
        except Exception as e:
            self.logger.error(f"Error in comprehensive CSV parsing {file_path}: {str(e)}")
            # Create empty dataframe as final fallback
            return pd.DataFrame(), []
    
    def _validate_csv_parsing(self, df: pd.DataFrame) -> bool:
        """
        Validate if CSV parsing was successful by checking data characteristics
        
        Args:
            df: DataFrame to validate
            
        Returns:
            bool: True if parsing appears successful
        """
        if df is None or df.empty:
            return False
        
        # Check for reasonable dimensions
        if len(df.columns) == 1 and len(df) == 1:
            return False
        
        # Check if we have reasonable data distribution
        non_empty_cells = 0
        total_cells = len(df) * len(df.columns)
        
        for col in df.columns:
            non_empty_cells += (df[col].astype(str).str.strip() != '').sum()
        
        # At least 10% of cells should have data
        if total_cells > 0 and (non_empty_cells / total_cells) < 0.1:
            return False
        
        return True
    
    def _fallback_csv_parsing(self, file_path: str) -> pd.DataFrame:
        """
        Fallback CSV parsing method for problematic files
        
        Args:
            file_path: Path to the CSV file
            
        Returns:
            pd.DataFrame: Parsed dataframe or empty if failed
        """
        try:
            # Try reading as raw text and parsing manually
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                lines = f.readlines()
            
            if not lines:
                return pd.DataFrame()
            
            # Try to detect delimiter from first few lines
            sample_lines = lines[:5]
            potential_delimiters = [',', ';', '\t', '|']
            delimiter_counts = {}
            
            for delimiter in potential_delimiters:
                count = sum(line.count(delimiter) for line in sample_lines)
                delimiter_counts[delimiter] = count
            
            # Use the delimiter that appears most frequently
            best_delimiter = max(delimiter_counts, key=delimiter_counts.get) if delimiter_counts else ','
            
            # Parse manually
            data = []
            for line in lines:
                row = [cell.strip().strip('"').strip("'") for cell in line.split(best_delimiter)]
                data.append(row)
            
            if data:
                # Find max columns
                max_cols = max(len(row) for row in data)
                # Pad rows to same length
                padded_data = [row + [''] * (max_cols - len(row)) for row in data]
                
                df = pd.DataFrame(padded_data)
                return df
            
        except Exception as e:
            self.logger.error(f"Fallback CSV parsing failed: {str(e)}")
        
        return pd.DataFrame()
    
    def _clean_csv_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply comprehensive cleaning to CSV data
        
        Args:
            df: DataFrame to clean
            
        Returns:
            pd.DataFrame: Cleaned DataFrame
        """
        if df.empty:
            return df
        
        df_cleaned = df.copy()
        
        # Clean all string values
        for col in df_cleaned.columns:
            df_cleaned[col] = df_cleaned[col].astype(str)
            # Remove quotes and clean whitespace
            df_cleaned[col] = df_cleaned[col].str.strip().str.strip('"').str.strip("'").str.strip()
            # Replace various representations of missing data
            df_cleaned[col] = df_cleaned[col].replace(['nan', 'NaN', 'NULL', 'null', 'N/A', 'n/a', '#N/A', '#NULL!'], '')
        
        # Remove completely empty rows and columns
        # Remove empty rows
        non_empty_rows = ~(df_cleaned == '').all(axis=1)
        if non_empty_rows.any():
            df_cleaned = df_cleaned[non_empty_rows].reset_index(drop=True)
        
        # Remove empty columns
        non_empty_cols = ~(df_cleaned == '').all(axis=0)
        if non_empty_cols.any():
            df_cleaned = df_cleaned.loc[:, non_empty_cols]
            # Reset column names to be sequential
            df_cleaned.columns = range(len(df_cleaned.columns))
        
        return df_cleaned
    
    def _fallback_excel_processing_enhanced(self, file_path: str) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
        """
        Enhanced fallback Excel processing using pandas when openpyxl fails
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Tuple[pd.DataFrame, List[Dict]]: Main dataframe and list of sheets
        """
        try:
            self.logger.info("Attempting enhanced fallback Excel processing...")
            
            # Read all sheets with multiple approaches
            excel_file = pd.ExcelFile(file_path)
            sheets_data = []
            main_df = None
            
            for sheet_name in excel_file.sheet_names:
                self.logger.debug(f"Processing sheet in fallback mode: {sheet_name}")
                
                # Try multiple reading approaches
                df = None
                
                # Approach 1: Read without headers, everything as string
                try:
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        header=None,
                        dtype=str,
                        keep_default_na=False,
                        na_filter=False
                    )
                except Exception as e:
                    self.logger.debug(f"Approach 1 failed for {sheet_name}: {str(e)}")
                
                # Approach 2: Skip potential header rows
                if df is None or df.empty:
                    try:
                        df = pd.read_excel(
                            excel_file,
                            sheet_name=sheet_name,
                            header=None,
                            skiprows=1,  # Skip first row in case it's problematic
                            dtype=str,
                            keep_default_na=False,
                            na_filter=False
                        )
                    except Exception as e:
                        self.logger.debug(f"Approach 2 failed for {sheet_name}: {str(e)}")
                
                # Approach 3: Use default pandas behavior as last resort
                if df is None or df.empty:
                    try:
                        df = pd.read_excel(
                            excel_file,
                            sheet_name=sheet_name,
                            dtype=str
                        )
                        # Reset index to treat header as data
                        if not df.empty:
                            # Insert the original column names as the first row
                            header_row = pd.DataFrame([df.columns.tolist()], columns=df.columns)
                            df = pd.concat([header_row, df], ignore_index=True)
                            # Reset column names to numeric
                            df.columns = range(len(df.columns))
                    except Exception as e:
                        self.logger.debug(f"Approach 3 failed for {sheet_name}: {str(e)}")
                
                if df is not None and not df.empty:
                    # Apply comprehensive cleaning
                    df = self._clean_fallback_excel_data(df)
                    
                    sheet_info = {
                        'name': sheet_name,
                        'data': df
                    }
                    sheets_data.append(sheet_info)
                    
                    if main_df is None:
                        main_df = df
            
            excel_file.close()
            
            if main_df is None:
                self.logger.warning("Enhanced fallback processing found no data")
                main_df = pd.DataFrame()
            
            return main_df, sheets_data
            
        except Exception as e:
            self.logger.error(f"Enhanced fallback Excel processing failed for {file_path}: {str(e)}")
            return pd.DataFrame(), []
    
    def _clean_fallback_excel_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean data from fallback Excel processing
        
        Args:
            df: DataFrame to clean
            
        Returns:
            pd.DataFrame: Cleaned DataFrame
        """
        if df.empty:
            return df
        
        df_cleaned = df.copy()
        
        # Convert all to string and clean
        for col in df_cleaned.columns:
            df_cleaned[col] = df_cleaned[col].astype(str)
            # Clean various NaN representations
            df_cleaned[col] = df_cleaned[col].replace(['nan', 'NaN', 'NaT', '<NA>', 'None'], '')
            # Clean whitespace
            df_cleaned[col] = df_cleaned[col].str.strip()
        
        # Remove completely empty rows and columns
        # Remove rows where all values are empty
        non_empty_rows = ~(df_cleaned == '').all(axis=1)
        if non_empty_rows.any():
            df_cleaned = df_cleaned[non_empty_rows].reset_index(drop=True)
        
        # Remove columns where all values are empty
        non_empty_cols = ~(df_cleaned == '').all(axis=0)
        if non_empty_cols.any():
            df_cleaned = df_cleaned.loc[:, non_empty_cols]
            # Reset column names
            df_cleaned.columns = range(len(df_cleaned.columns))
        
        return df_cleaned
    
    def save_as_clean_csv(self, df: pd.DataFrame, output_path: str) -> None:
        """
        Save DataFrame as clean CSV with enhanced options
        
        Args:
            df: DataFrame to save
            output_path: Path where to save the CSV
        """
        try:
            self.logger.info(f"Saving clean CSV to: {output_path}")
            
            # Ensure the output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Save with enhanced CSV settings
            df.to_csv(
                output_path,
                index=False,
                header=False,  # No headers since we treat everything as data
                encoding='utf-8',
                quoting=csv.QUOTE_MINIMAL,
                quotechar='"',
                lineterminator='\n',
                escapechar=None,
                doublequote=True
            )
            
            self.logger.info(f"Successfully saved {len(df)} rows and {len(df.columns)} columns to CSV")
            
        except Exception as e:
            self.logger.error(f"Error saving clean CSV to {output_path}: {str(e)}")
            raise
    
    def save_as_original_format(self, df: pd.DataFrame, output_path: str, original_file_type: str) -> None:
        """
        Save DataFrame in the original file format
        
        Args:
            df: DataFrame to save
            output_path: Path where to save the file
            original_file_type: Original file extension (.xlsx, .xls, .csv)
        """
        try:
            self.logger.info(f"Saving file in original format: {original_file_type}")
            
            # Ensure the output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            if original_file_type.lower() == '.csv':
                # Save as CSV
                df.to_csv(
                    output_path,
                    index=False,
                    header=False,  # No headers since we treat everything as data
                    encoding='utf-8',
                    quoting=csv.QUOTE_MINIMAL,
                    quotechar='"',
                    lineterminator='\n',
                    escapechar=None,
                    doublequote=True
                )
            elif original_file_type.lower() in ['.xlsx', '.xls']:
                # Save as Excel
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    df.to_excel(
                        writer, 
                        sheet_name='Sheet1', 
                        index=False, 
                        header=False  # No headers since we treat everything as data
                    )
            else:
                # Fallback to CSV if unknown format
                self.logger.warning(f"Unknown file type {original_file_type}, saving as CSV")
                df.to_csv(
                    output_path,
                    index=False,
                    header=False,
                    encoding='utf-8',
                    quoting=csv.QUOTE_MINIMAL
                )
            
            self.logger.info(f"Successfully saved {len(df)} rows and {len(df.columns)} columns as {original_file_type}")
            
        except Exception as e:
            self.logger.error(f"Error saving file as {original_file_type} to {output_path}: {str(e)}")
            raise
    
    def get_parsing_summary(self, df: pd.DataFrame, sheets: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Generate a summary of the parsing results
        
        Args:
            df: Main dataframe
            sheets: List of sheet information
            
        Returns:
            Dict: Summary information
        """
        summary = {
            'total_sheets': len(sheets),
            'main_sheet_rows': len(df) if df is not None else 0,
            'main_sheet_cols': len(df.columns) if df is not None and not df.empty else 0,
            'total_cells': 0,
            'non_empty_cells': 0,
            'sheets_info': []
        }
        
        for sheet_info in sheets:
            sheet_df = sheet_info['data']
            if sheet_df is not None and not sheet_df.empty:
                rows = len(sheet_df)
                cols = len(sheet_df.columns)
                total_cells = rows * cols
                non_empty = 0
                
                for col in sheet_df.columns:
                    non_empty += (sheet_df[col].astype(str).str.strip() != '').sum()
                
                sheet_summary = {
                    'name': sheet_info['name'],
                    'rows': rows,
                    'columns': cols,
                    'total_cells': total_cells,
                    'non_empty_cells': non_empty,
                    'data_density': (non_empty / total_cells * 100) if total_cells > 0 else 0
                }
                
                summary['sheets_info'].append(sheet_summary)
                summary['total_cells'] += total_cells
                summary['non_empty_cells'] += non_empty
        
        summary['overall_data_density'] = (
            summary['non_empty_cells'] / summary['total_cells'] * 100 
            if summary['total_cells'] > 0 else 0
        )
        
        return summary

    # Legacy methods for backward compatibility
    @staticmethod
    def parse_to_json(file_path: str) -> Tuple[str, Dict[str, Any]]:
        """
        Parse a spreadsheet file to JSON
        Note: This method now expects files to be preprocessed into clean CSV format

        Args:
            file_path: Path to the spreadsheet file

        Returns:
            Tuple[str, Dict]: JSON representation and metadata
        """
        file_ext = os.path.splitext(file_path)[1].lower()
        
        # Read file based on extension
        # All files are now treated as clean data (no headers, all strings)
        if file_ext in ['.xlsx', '.xls']:
            # For Excel files, read without headers since data is preprocessed
            df = pd.read_excel(file_path, header=None, dtype=str)
        elif file_ext == '.csv':
            # For CSV files, read without headers since data is preprocessed
            df = pd.read_csv(file_path, header=None, dtype=str)
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")
        
        # Clean any remaining NaN values
        df = df.fillna('')
        
        # Convert to JSON
        data = df.to_json(orient='records')
        
        # Prepare metadata (no column names since we treat everything as data)
        metadata = {
            'columns': [f"Column_{i}" for i in range(len(df.columns))],
            'rows': len(df),
            'file_type': file_ext[1:]  # Remove the dot
        }
        
        return data, metadata
    
    @staticmethod
    def detect_file_format(file_path: str) -> str:
        """
        Detect the format of a spreadsheet file

        Args:
            file_path: Path to the spreadsheet file

        Returns:
            str: Detected format (excel, csv)
        """
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext in ['.xlsx', '.xls']:
            return 'excel'
        elif file_ext == '.csv':
            return 'csv'
        else:
            raise ValueError(f"Unsupported file format: {file_ext}")
    
    @staticmethod
    def parse_from_pandas(df: pd.DataFrame) -> str:
        """
        Convert a pandas DataFrame to JSON string

        Args:
            df: Pandas DataFrame

        Returns:
            str: JSON representation of the DataFrame
        """
        # Ensure all data is string type and clean
        df_clean = df.copy()
        for col in df_clean.columns:
            df_clean[col] = df_clean[col].astype(str).fillna('')
        
        return df_clean.to_json(orient='records')
